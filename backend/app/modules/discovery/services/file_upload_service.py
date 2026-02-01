"""File upload service for S3 storage and CSV/XLSX parsing.

Provides the FileUploadService which handles:
- File upload to S3 with unique key generation
- CSV and XLSX file parsing with schema detection
- Column mapping suggestions based on common patterns
- Unique value extraction for role analysis
- File validation and size limits
"""

import csv
import io
from collections import Counter
from typing import Any, Protocol
from uuid import UUID, uuid4


class S3ClientProtocol(Protocol):
    """Protocol for S3 client interface."""

    async def upload_fileobj(
        self,
        fileobj: io.BytesIO,
        bucket: str,
        key: str,
        **kwargs: Any,
    ) -> None:
        """Upload file object to S3."""
        ...

    async def delete_object(self, bucket: str, key: str) -> None:
        """Delete object from S3."""
        ...

    async def generate_presigned_url(
        self,
        client_method: str,
        params: dict[str, Any],
        expires_in: int,
    ) -> str:
        """Generate presigned URL for S3 object."""
        ...


class UploadRepositoryProtocol(Protocol):
    """Protocol for upload repository interface."""

    async def create(
        self,
        session_id: UUID,
        file_name: str,
        file_url: str,
        row_count: int,
        column_mappings: dict | None = None,
        detected_schema: dict | None = None,
    ) -> Any:
        """Create upload record."""
        ...

    async def get_by_id(self, upload_id: UUID) -> Any | None:
        """Get upload by ID."""
        ...

    async def delete(self, upload_id: UUID) -> bool:
        """Delete upload record."""
        ...


class FileUploadService:
    """Service for handling file uploads and parsing.

    This service provides functionality for:
    - Uploading files to S3 with organized key structure
    - Parsing CSV and XLSX files to detect schema
    - Suggesting column mappings based on common naming patterns
    - Extracting unique values from specific columns
    - Validating file size and format

    Attributes:
        s3_client: Client for S3 operations.
        upload_repo: Repository for upload record management.
        bucket_name: Name of the S3 bucket for uploads.
        max_file_size: Maximum allowed file size in bytes (100MB default).
    """

    # Maximum file size: 100MB
    MAX_FILE_SIZE = 100 * 1024 * 1024

    # Supported file extensions
    SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls"}

    # Sample row limit for preview
    SAMPLE_ROW_LIMIT = 5

    # Row limit for type detection sampling
    TYPE_DETECTION_ROW_LIMIT = 100

    # Column name patterns for role mapping suggestions
    ROLE_PATTERNS = [
        "job title",
        "jobtitle",
        "job_title",
        "position",
        "title",
        "job",
        "occupation",
        "role",
    ]

    # Column name patterns for department mapping suggestions
    DEPARTMENT_PATTERNS = [
        "department",
        "dept",
        "division",
        "team",
        "group",
        "unit",
    ]

    def __init__(
        self,
        s3_client: S3ClientProtocol,
        upload_repo: UploadRepositoryProtocol,
        bucket_name: str,
    ) -> None:
        """Initialize the service with dependencies.

        Args:
            s3_client: S3 client for file operations.
            upload_repo: Repository for upload records.
            bucket_name: Name of the S3 bucket.
        """
        self.s3_client = s3_client
        self.upload_repo = upload_repo
        self.bucket_name = bucket_name
        self.max_file_size = self.MAX_FILE_SIZE

    async def upload_file(
        self,
        session_id: UUID,
        file_name: str,
        file_content: bytes,
    ) -> dict[str, Any]:
        """Upload a file to S3.

        Generates a unique S3 key with session ID prefix for organization
        and uploads the file content.

        Args:
            session_id: UUID of the discovery session.
            file_name: Original name of the file.
            file_content: Binary content of the file.

        Returns:
            Dictionary containing:
            - file_url: S3 URL of the uploaded file
            - s3_key: The S3 key used for storage

        Raises:
            ValueError: If file fails validation (size or type).
            RuntimeError: If S3 upload fails.
        """
        # Validate file before upload
        await self.validate_file(file_name, file_content)

        # Generate unique S3 key with session prefix
        unique_id = uuid4()
        s3_key = f"sessions/{session_id}/{unique_id}_{file_name}"

        # Create file object for upload
        file_obj = io.BytesIO(file_content)

        # Upload to S3
        try:
            await self.s3_client.upload_fileobj(
                fileobj=file_obj,
                bucket=self.bucket_name,
                key=s3_key,
            )
        except Exception as e:
            raise RuntimeError(f"Failed to upload file to S3: {e}") from e

        # Construct S3 URL
        file_url = f"s3://{self.bucket_name}/{s3_key}"

        return {
            "file_url": file_url,
            "s3_key": s3_key,
        }

    async def parse_file(
        self,
        file_name: str,
        file_content: bytes,
    ) -> dict[str, Any]:
        """Parse a file and detect its schema.

        Supports CSV and XLSX file formats. Detects columns, data types,
        and extracts sample values for preview.

        Args:
            file_name: Name of the file (used to determine format).
            file_content: Binary content of the file.

        Returns:
            Dictionary containing:
            - row_count: Number of data rows (excluding header)
            - detected_schema: Schema information including columns, types, sample_values

        Raises:
            ValueError: If file is empty or has unsupported format.
        """
        # Validate file is not empty
        if not file_content:
            raise ValueError("File is empty")

        # Determine file type from extension
        extension = self._get_file_extension(file_name)

        if extension not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(f"Unsupported file type: {extension}")

        # Route to appropriate parser
        if extension == ".csv":
            return await self._parse_csv(file_content)
        else:
            # XLSX/XLS parsing
            return await self._parse_xlsx(file_content)

    async def _parse_csv(self, file_content: bytes) -> dict[str, Any]:
        """Parse CSV file content.

        Args:
            file_content: Binary content of the CSV file.

        Returns:
            Parsed schema and row count information.
        """
        # Decode content
        try:
            text_content = file_content.decode("utf-8")
        except UnicodeDecodeError:
            # Try alternative encoding
            text_content = file_content.decode("latin-1")

        # Parse CSV
        reader = csv.reader(io.StringIO(text_content))
        rows = list(reader)

        if not rows:
            raise ValueError("File is empty")

        # Extract header and data rows
        header = rows[0]
        data_rows = rows[1:]

        # Detect column types from data
        column_types = self._detect_column_types(header, data_rows)

        # Extract sample values (up to SAMPLE_ROW_LIMIT rows)
        sample_values = data_rows[:self.SAMPLE_ROW_LIMIT]

        return {
            "row_count": len(data_rows),
            "detected_schema": {
                "columns": header,
                "types": column_types,
                "sample_values": sample_values,
            },
        }

    async def _parse_xlsx(self, file_content: bytes) -> dict[str, Any]:
        """Parse XLSX/XLS file content.

        This method requires openpyxl or xlrd library for full functionality.
        Currently provides a basic implementation that can be extended.

        Args:
            file_content: Binary content of the Excel file.

        Returns:
            Parsed schema and row count information.
        """
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
            sheet = workbook.active

            if sheet is None:
                raise ValueError("No active sheet in workbook")

            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                raise ValueError("File is empty")

            # Extract header and data rows
            header = [str(cell) if cell is not None else "" for cell in rows[0]]
            data_rows = [
                [str(cell) if cell is not None else "" for cell in row]
                for row in rows[1:]
            ]

            # Detect column types
            column_types = self._detect_column_types(header, data_rows)

            # Extract sample values
            sample_values = data_rows[:self.SAMPLE_ROW_LIMIT]

            workbook.close()

            return {
                "row_count": len(data_rows),
                "detected_schema": {
                    "columns": header,
                    "types": column_types,
                    "sample_values": sample_values,
                },
            }

        except ImportError:
            # openpyxl not installed, return minimal schema
            return {
                "row_count": 0,
                "detected_schema": {
                    "columns": [],
                    "types": [],
                    "sample_values": [],
                },
            }

    def _detect_column_types(
        self,
        header: list[str],
        data_rows: list[list[str]],
    ) -> list[str]:
        """Detect data types for each column.

        Analyzes sample data to determine likely type:
        - integer: All numeric values without decimals
        - float: Numeric values with decimals
        - boolean: true/false values
        - string: All other values

        Args:
            header: List of column names.
            data_rows: List of data rows.

        Returns:
            List of detected type names for each column.
        """
        num_columns = len(header)
        types = []

        for col_idx in range(num_columns):
            # Get non-empty values for this column
            values = []
            for row in data_rows[:self.TYPE_DETECTION_ROW_LIMIT]:  # Sample rows for type detection
                if col_idx < len(row) and row[col_idx]:
                    values.append(str(row[col_idx]).strip())

            if not values:
                types.append("string")
                continue

            # Check for integer
            if all(self._is_integer(v) for v in values):
                types.append("integer")
            # Check for float
            elif all(self._is_numeric(v) for v in values):
                types.append("float")
            # Check for boolean
            elif all(v.lower() in ("true", "false", "yes", "no", "1", "0") for v in values):
                types.append("boolean")
            else:
                types.append("string")

        return types

    def _is_integer(self, value: str) -> bool:
        """Check if value represents an integer."""
        try:
            int(value)
            return "." not in value
        except ValueError:
            return False

    def _is_numeric(self, value: str) -> bool:
        """Check if value represents a number."""
        try:
            float(value.replace(",", ""))
            return True
        except ValueError:
            return False

    def _get_file_extension(self, file_name: str) -> str:
        """Extract file extension from filename."""
        if "." not in file_name:
            return ""
        return "." + file_name.rsplit(".", 1)[-1].lower()

    async def suggest_column_mappings(
        self,
        detected_schema: dict[str, Any],
    ) -> dict[str, str]:
        """Suggest column mappings based on column names.

        Analyzes column names against known patterns for role and department
        fields to suggest automatic mappings.

        Args:
            detected_schema: Schema information from parse_file.

        Returns:
            Dictionary mapping field names to suggested column names:
            - role: Suggested column for job role/title
            - department: Suggested column for department/team
        """
        columns = detected_schema.get("columns", [])
        suggestions: dict[str, str] = {}

        # Find role column
        for col in columns:
            col_lower = col.lower().strip()
            for pattern in self.ROLE_PATTERNS:
                if pattern in col_lower or col_lower == pattern:
                    suggestions["role"] = col
                    break
            if "role" in suggestions:
                break

        # Find department column
        for col in columns:
            col_lower = col.lower().strip()
            for pattern in self.DEPARTMENT_PATTERNS:
                if pattern in col_lower or col_lower == pattern:
                    suggestions["department"] = col
                    break
            if "department" in suggestions:
                break

        return suggestions

    async def upload_and_register(
        self,
        session_id: UUID,
        file_name: str,
        file_content: bytes,
    ) -> dict[str, Any]:
        """Upload file, parse it, and register in database.

        Combines upload, parsing, and registration into a single operation
        for convenience.

        Args:
            session_id: UUID of the discovery session.
            file_name: Name of the file.
            file_content: Binary content of the file.

        Returns:
            Dictionary containing:
            - upload_id: UUID of the created upload record
            - file_url: S3 URL of the uploaded file
            - parsed_data: Parsed schema and row count information
        """
        # Upload to S3
        upload_result = await self.upload_file(
            session_id=session_id,
            file_name=file_name,
            file_content=file_content,
        )

        # Parse the file
        parsed_data = await self.parse_file(
            file_name=file_name,
            file_content=file_content,
        )

        # Register in database with rollback on failure
        try:
            upload_record = await self.upload_repo.create(
                session_id=session_id,
                file_name=file_name,
                file_url=upload_result["file_url"],
                row_count=parsed_data["row_count"],
                detected_schema=parsed_data["detected_schema"],
            )
        except Exception as e:
            # Clean up orphaned S3 file
            try:
                await self.s3_client.delete_object(
                    bucket=self.bucket_name,
                    key=upload_result["s3_key"],
                )
            except Exception:
                pass  # Best effort cleanup
            raise

        return {
            "upload_id": upload_record.id,
            "file_url": upload_result["file_url"],
            "parsed_data": parsed_data,
        }

    async def delete_upload(self, upload_id: UUID) -> None:
        """Delete an upload from S3 and database.

        Retrieves the upload record, deletes the file from S3,
        then removes the database record.

        Args:
            upload_id: UUID of the upload to delete.

        Raises:
            ValueError: If upload not found.
        """
        # Get upload record
        upload = await self.upload_repo.get_by_id(upload_id)
        if upload is None:
            raise ValueError(f"Upload not found: {upload_id}")

        # Extract S3 key from URL
        file_url = upload.file_url
        s3_key = self._extract_s3_key(file_url)

        # Delete from S3
        try:
            await self.s3_client.delete_object(
                bucket=self.bucket_name,
                key=s3_key,
            )
        except Exception as e:
            raise RuntimeError(f"Failed to delete file from S3: {e}") from e

        # Delete from database
        await self.upload_repo.delete(upload_id)

    def _extract_s3_key(self, file_url: str) -> str:
        """Extract S3 key from S3 URL.

        Args:
            file_url: S3 URL (e.g., s3://bucket/path/to/file)

        Returns:
            S3 key (path portion of URL).
        """
        # Handle s3:// URLs
        if file_url.startswith("s3://"):
            # Remove s3://bucket_name/ prefix
            parts = file_url[5:].split("/", 1)
            if len(parts) > 1:
                return parts[1]
        return file_url

    async def get_download_url(
        self,
        file_url: str,
        expires_in: int = 3600,
    ) -> str:
        """Generate a presigned download URL for a file.

        Args:
            file_url: S3 URL of the file.
            expires_in: URL expiration time in seconds (default 1 hour).

        Returns:
            Presigned HTTPS URL for downloading the file.
        """
        s3_key = self._extract_s3_key(file_url)

        try:
            presigned_url = await self.s3_client.generate_presigned_url(
                client_method="get_object",
                params={
                    "Bucket": self.bucket_name,
                    "Key": s3_key,
                },
                expires_in=expires_in,
            )
        except Exception as e:
            raise RuntimeError(f"Failed to generate presigned URL: {e}") from e

        return presigned_url

    async def extract_unique_values(
        self,
        file_content: bytes,
        file_name: str,
        column_name: str,
    ) -> dict[str, int]:
        """Extract unique values from a specific column with counts.

        Parses the file and counts occurrences of each unique value
        in the specified column.

        Args:
            file_content: Binary content of the file.
            file_name: Name of the file.
            column_name: Name of the column to extract values from.

        Returns:
            Dictionary mapping unique values to their occurrence counts.

        Raises:
            ValueError: If column name does not exist in the file.
        """
        # Get file extension to determine parsing strategy
        extension = self._get_file_extension(file_name)

        # Parse once and get both columns and all rows
        if extension == ".csv":
            columns, all_rows = await self._get_csv_columns_and_rows(file_content)
        else:
            columns, all_rows = await self._get_xlsx_columns_and_rows(file_content)

        # Find column index
        if column_name not in columns:
            raise ValueError(f"Column '{column_name}' not found in file")

        col_idx = columns.index(column_name)

        # Extract values from the column
        values = []
        for row in all_rows:
            if col_idx < len(row) and row[col_idx]:
                values.append(str(row[col_idx]).strip())

        # Count unique values
        return dict(Counter(values))

    async def _get_csv_columns_and_rows(
        self, file_content: bytes
    ) -> tuple[list[str], list[list[str]]]:
        """Get columns and all data rows from CSV content in a single parse."""
        try:
            text_content = file_content.decode("utf-8")
        except UnicodeDecodeError:
            text_content = file_content.decode("latin-1")

        reader = csv.reader(io.StringIO(text_content))
        rows = list(reader)

        if not rows:
            return [], []

        return rows[0], rows[1:]  # columns, data rows

    async def _get_xlsx_columns_and_rows(
        self, file_content: bytes
    ) -> tuple[list[str], list[list[str]]]:
        """Get columns and all data rows from XLSX content in a single parse."""
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
            sheet = workbook.active

            if sheet is None:
                return [], []

            rows = list(sheet.iter_rows(values_only=True))
            workbook.close()

            if not rows:
                return [], []

            # Convert to strings
            columns = [str(cell) if cell is not None else "" for cell in rows[0]]
            data_rows = [
                [str(cell) if cell is not None else "" for cell in row]
                for row in rows[1:]
            ]

            return columns, data_rows

        except ImportError:
            return [], []

    async def validate_file(
        self,
        file_name: str,
        file_content: bytes,
    ) -> dict[str, Any]:
        """Validate a file before processing.

        Checks:
        - File size is within limits (100MB)
        - File extension is supported

        Args:
            file_name: Name of the file.
            file_content: Binary content of the file.

        Returns:
            Dictionary containing validation status and any warnings.

        Raises:
            ValueError: If file fails validation.
        """
        warnings = []

        # Check file size
        file_size = len(file_content)
        if file_size > self.max_file_size:
            raise ValueError(
                f"File size exceeds maximum allowed size of "
                f"{self.max_file_size // (1024 * 1024)}MB"
            )

        # Check file extension
        extension = self._get_file_extension(file_name)
        if extension not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(f"Unsupported file type: {extension}")

        return {
            "valid": True,
            "file_size": file_size,
            "extension": extension,
            "warnings": warnings,
        }
